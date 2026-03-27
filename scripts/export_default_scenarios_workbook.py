from __future__ import annotations

import json
import math
import re
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SCENARIO_FILE = ROOT / "simulator" / "scenario.js"
REGISTRY_FILE = ROOT / "simulator" / "registry.js"
OUTPUT_FILE = ROOT / "docs" / "default_scenarios_metrics.xlsx"
FALLBACK_USAGE_RATE = 65


def read_utf8(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def load_default_scenarios() -> list[dict]:
    node_script = f"""
import {{ defaultScenarios }} from {json.dumps(SCENARIO_FILE.as_uri())};
console.log(JSON.stringify(defaultScenarios()));
"""
    completed = subprocess.run(
        ["node", "--input-type=module", "-"],
        input=node_script,
        text=True,
        capture_output=True,
        check=True,
        cwd=ROOT,
        encoding="utf-8",
    )
    return json.loads(completed.stdout)


def parse_registry_meta() -> dict[str, dict[str, str]]:
    text = read_utf8(REGISTRY_FILE)
    pattern = re.compile(r"^  ([A-Za-z0-9_]+): createMeta\(\{(.*?)^  \}\),", re.MULTILINE | re.DOTALL)
    meta: dict[str, dict[str, str]] = {}

    for key, body in pattern.findall(text):
        item: dict[str, str] = {}
        for field in ("label", "unit", "tip", "prefix", "suffix", "sticker"):
            match = re.search(rf"{field}:\s*(['\"])(.*?)\1", body, re.DOTALL)
            if match:
                item[field] = match.group(2)
        meta[key] = item

    return meta


def average(values: list[float], fallback: float = 0) -> float:
    if not values:
        return fallback
    return sum(values) / len(values)


def get_usage_values(scenario: dict, region_id: str) -> list[float]:
    return scenario.get("usageByRegion", {}).get(region_id, [])


def get_average_usage_rate_by_region(scenario: dict, region: dict) -> float:
    return average(get_usage_values(scenario, region["id"]), 100)


def get_total_round_limit(scenario: dict) -> float:
    return max(sum(float(region.get("roundLimitByRegion", 0) or 0) for region in scenario["regions"]), 1)


def get_year_usage_rate(scenario: dict, region: dict, year_index: int) -> float:
    values = scenario.get("usageByRegion", {}).get(region["id"], [])
    return float(values[year_index] if year_index < len(values) else FALLBACK_USAGE_RATE) / 100


def year_churn_rate(scenario: dict, year_index: int) -> float:
    churn_values = scenario.get("churnByYear", [])
    base = scenario.get("yearlyChurnRate", 0)
    raw = churn_values[year_index] if year_index < len(churn_values) else base
    return float(raw or 0) / 100


def year_refund_amount(scenario: dict, year_index: int) -> float:
    refund_values = scenario.get("refundAmountByYear", [])
    base = scenario.get("yearlyRefundAmount", 0)
    raw = refund_values[year_index] if year_index < len(refund_values) else base
    return float(raw or 0)


def year_surviving_end(prev_surviving: float, churn_rate: float) -> float:
    return prev_surviving - prev_surviving * churn_rate


def year_refund_cost(prev_surviving: float, churn_rate: float, refund_amount: float) -> float:
    return prev_surviving * churn_rate * refund_amount


def member_saving_by_region(region: dict) -> float:
    return (region["publicPriceByRegion"] - region["memberPriceByRegion"]) * region["membersByRegion"]


def companion_saving_by_region(region: dict) -> float:
    return (region["publicPriceByRegion"] - region["companionPriceByRegion"]) * region["companionsByRegion"]


def member_profit_by_region(region: dict) -> float:
    return (region["memberPriceByRegion"] - region["courseCostByRegion"]) * region["membersByRegion"]


def companion_profit_by_region(region: dict) -> float:
    return (region["companionPriceByRegion"] - region["courseCostByRegion"]) * region["companionsByRegion"]


def margin_per_round_by_region(region: dict) -> float:
    return member_profit_by_region(region) + companion_profit_by_region(region)


def actual_rounds_by_region(scenario: dict, region: dict) -> float:
    return region["roundLimitByRegion"] * (get_average_usage_rate_by_region(scenario, region) / 100)


def total_actual_rounds(scenario: dict) -> float:
    return sum(actual_rounds_by_region(scenario, region) for region in scenario["regions"])


def weighted_average_by_actual_rounds(scenario: dict, pick_value) -> float:
    total_rounds = total_actual_rounds(scenario)
    if total_rounds == 0:
        return 0
    return sum(
        (actual_rounds_by_region(scenario, region) / total_rounds) * pick_value(region)
        for region in scenario["regions"]
    )


def total_annual_rounding(scenario: dict) -> float:
    return sum(margin_per_round_by_region(region) * actual_rounds_by_region(scenario, region) for region in scenario["regions"])


def saving_per_round_by_region(region: dict) -> float:
    return member_saving_by_region(region) + companion_saving_by_region(region)


def total_annual_saving(scenario: dict) -> float:
    return sum(saving_per_round_by_region(region) * actual_rounds_by_region(scenario, region) for region in scenario["regions"])


def sum_expected_by_year(scenario: dict, pick_value) -> float:
    total = 0.0
    previous_surviving = 1.0
    for year_index in range(scenario["contractYears"]):
        total += pick_value(year_index, previous_surviving)
        previous_surviving = year_surviving_end(previous_surviving, year_churn_rate(scenario, year_index))
    return total


def build_years(scenario: dict) -> list[dict]:
    years: list[dict] = []
    previous_surviving = 1.0
    cumulative_ltv = 0.0

    for year_index in range(scenario["contractYears"]):
        churn_rate = year_churn_rate(scenario, year_index)
        refund_amount = year_refund_amount(scenario, year_index)
        enroll_fee_this_year = scenario["salePrice"] if year_index == 0 else 0
        surviving = year_surviving_end(previous_surviving, churn_rate)
        refund_cost = year_refund_cost(previous_surviving, churn_rate, refund_amount)

        year_usage_weighted = (
            sum(
                get_year_usage_rate(scenario, region, year_index)
                * (region["roundLimitByRegion"] / get_total_round_limit(scenario))
                for region in scenario["regions"]
            )
            * 100
        )
        year_actual_rounds_expected = sum(
            region["roundLimitByRegion"] * get_year_usage_rate(scenario, region, year_index) * previous_surviving
            for region in scenario["regions"]
        )
        year_member_revenue_expected = sum(
            region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            * region["memberPriceByRegion"]
            * region["membersByRegion"]
            * previous_surviving
            for region in scenario["regions"]
        )
        year_companion_revenue_expected = sum(
            region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            * region["companionPriceByRegion"]
            * region["companionsByRegion"]
            * previous_surviving
            for region in scenario["regions"]
        )
        year_rounding_cost_expected = sum(
            region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            * region["courseCostByRegion"]
            * (region["membersByRegion"] + region["companionsByRegion"])
            * previous_surviving
            for region in scenario["regions"]
        )
        year_rounding_revenue_expected = year_member_revenue_expected + year_companion_revenue_expected
        year_rounding_revenue = sum(
            margin_per_round_by_region(region)
            * region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            for region in scenario["regions"]
        )
        year_variable_profit_expected = (
            year_member_revenue_expected
            + year_companion_revenue_expected
            + scenario["annualFixedProfit"] * previous_surviving
            - year_rounding_cost_expected
        )
        year_cash_flow_expected = (
            enroll_fee_this_year
            + scenario["annualFee"] * previous_surviving
            - refund_cost
            + year_variable_profit_expected
        )
        year_revenue = (
            enroll_fee_this_year
            + scenario["annualFee"] * previous_surviving
            + scenario["annualFixedProfit"] * previous_surviving
            + year_rounding_revenue * previous_surviving
            - refund_cost
        )
        year_fixed_profit = (
            enroll_fee_this_year
            + scenario["annualFee"] * previous_surviving
            + scenario["annualFixedProfit"] * previous_surviving
            - refund_cost
        )
        cumulative_ltv += year_revenue

        years.append(
            {
                "yearByYear": year_index + 1,
                "yearLabelByYear": f"{year_index + 1}년차",
                "survivingStartByYear": round(previous_surviving * 100, 1),
                "survivingEndByYear": round(surviving * 100, 1),
                "usageByYear": round(year_usage_weighted, 1),
                "enrollFeeByYear": round(enroll_fee_this_year),
                "annualFeeRevenueByYear": round(scenario["annualFee"] * previous_surviving),
                "annualFixedProfitByYear": round(scenario["annualFixedProfit"] * previous_surviving),
                "actualRoundsExpectedByYear": round(year_actual_rounds_expected, 1),
                "memberRevenueExpectedByYear": round(year_member_revenue_expected),
                "companionRevenueExpectedByYear": round(year_companion_revenue_expected),
                "roundingRevenueExpectedByYear": round(year_rounding_revenue_expected),
                "roundingCostExpectedByYear": round(year_rounding_cost_expected),
                "variableProfitExpectedByYear": round(year_variable_profit_expected),
                "cashFlowExpectedByYear": round(year_cash_flow_expected),
                "revenueByYear": round(year_revenue),
                "cumulativeLTVByYear": round(cumulative_ltv),
                "refundCostByYear": round(refund_cost),
                "fixedProfitByYear": round(year_fixed_profit),
                "rawPrevSurviving": previous_surviving,
                "rawChurnRate": churn_rate,
                "rawRefundAmount": refund_amount,
            }
        )

        previous_surviving = surviving

    return years


def build_region_year_rows(scenario: dict, registry_meta: dict[str, dict[str, str]]) -> list[dict]:
    rows: list[dict] = []
    previous_surviving = 1.0

    custom_meta = {
        "yearlyUsageRateByRegion": registry_meta.get("yearlyUsageRateByRegion", {}),
        "yearActualRoundsExpectedByRegion": {
            "label": "연간 라운딩 횟수",
            "unit": "회",
            "tip": "구장별 연초 잔존율과 이용률 반영 기준 연간 라운딩 횟수",
        },
        "yearMemberRevenueExpectedByRegion": {
            "label": "연간 회원 판매가",
            "unit": "만원",
            "tip": "구장별 연초 잔존율 반영 기준 연간 회원 판매가",
        },
        "yearCompanionRevenueExpectedByRegion": {
            "label": "연간 동반 판매가",
            "unit": "만원",
            "tip": "구장별 연초 잔존율 반영 기준 연간 동반 판매가",
        },
        "yearRoundingRevenueExpectedByRegion": {
            "label": "라운딩 수입",
            "unit": "만원",
            "tip": "구장별 연간 회원 판매가와 연간 동반 판매가의 합",
        },
        "yearRoundingCostExpectedByRegion": {
            "label": "연간 매입가(B2B)",
            "unit": "만원",
            "tip": "구장별 연초 잔존율 반영 기준 연간 매입가(B2B)",
        },
    }

    for year_index in range(scenario["contractYears"]):
        for region in scenario["regions"]:
            usage_rate = get_year_usage_rate(scenario, region, year_index) * 100
            actual_rounds = region["roundLimitByRegion"] * (usage_rate / 100) * previous_surviving
            member_revenue = actual_rounds * region["memberPriceByRegion"] * region["membersByRegion"]
            companion_revenue = actual_rounds * region["companionPriceByRegion"] * region["companionsByRegion"]
            rounding_revenue = member_revenue + companion_revenue
            rounding_cost = actual_rounds * region["courseCostByRegion"] * (
                region["membersByRegion"] + region["companionsByRegion"]
            )

            for variable_key, value in [
                ("yearlyUsageRateByRegion", usage_rate),
                ("yearActualRoundsExpectedByRegion", actual_rounds),
                ("yearMemberRevenueExpectedByRegion", member_revenue),
                ("yearCompanionRevenueExpectedByRegion", companion_revenue),
                ("yearRoundingRevenueExpectedByRegion", rounding_revenue),
                ("yearRoundingCostExpectedByRegion", rounding_cost),
            ]:
                meta = custom_meta[variable_key]
                rows.append(
                    make_row(
                        category="구장-연도별 집계",
                        variable_key=variable_key,
                        region_name=region["name"],
                        year_index=year_index + 1,
                        value=value,
                        meta=meta,
                    )
                )

        previous_surviving = year_surviving_end(previous_surviving, year_churn_rate(scenario, year_index))

    return rows


def build_result_bundle(scenario: dict) -> dict:
    years = build_years(scenario)
    total_round_limit_value = get_total_round_limit(scenario)
    total_actual_rounds_value = total_actual_rounds(scenario)
    weighted_margin = weighted_average_by_actual_rounds(scenario, margin_per_round_by_region)
    weighted_saving = weighted_average_by_actual_rounds(scenario, saving_per_round_by_region)
    total_expected_annual_fee_revenue = sum_expected_by_year(
        scenario, lambda _year_index, previous_surviving: scenario["annualFee"] * previous_surviving
    )
    total_expected_positive_annual_fixed_profit = sum_expected_by_year(
        scenario,
        lambda _year_index, previous_surviving: max(float(scenario["annualFixedProfit"] or 0), 0) * previous_surviving,
    )
    total_expected_negative_annual_fixed_cost = sum_expected_by_year(
        scenario,
        lambda _year_index, previous_surviving: max(-float(scenario["annualFixedProfit"] or 0), 0) * previous_surviving,
    )
    total_expected_rounding_revenue = sum_expected_by_year(
        scenario,
        lambda year_index, previous_surviving: sum(
            (
                region["memberPriceByRegion"] * region["membersByRegion"]
                + region["companionPriceByRegion"] * region["companionsByRegion"]
            )
            * region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            * previous_surviving
            for region in scenario["regions"]
        ),
    )
    total_expected_rounding_cost = sum_expected_by_year(
        scenario,
        lambda year_index, previous_surviving: sum(
            region["courseCostByRegion"]
            * (region["membersByRegion"] + region["companionsByRegion"])
            * region["roundLimitByRegion"]
            * get_year_usage_rate(scenario, region, year_index)
            * previous_surviving
            for region in scenario["regions"]
        ),
    )
    total_expected_variable_revenue = (
        total_expected_annual_fee_revenue
        + total_expected_positive_annual_fixed_profit
        + total_expected_rounding_revenue
    )
    total_expected_variable_cost = total_expected_negative_annual_fixed_cost + total_expected_rounding_cost
    avg_annual_variable_revenue = total_expected_variable_revenue / scenario["contractYears"]
    avg_annual_variable_cost = total_expected_variable_cost / scenario["contractYears"]
    avg_annual_variable_profit = avg_annual_variable_revenue - avg_annual_variable_cost
    total_expected_refund_cost = sum_expected_by_year(
        scenario,
        lambda year_index, previous_surviving: previous_surviving
        * year_churn_rate(scenario, year_index)
        * year_refund_amount(scenario, year_index),
    )
    naive_fixed_profit_value = sum(year["fixedProfitByYear"] for year in years)
    ltv_value = years[-1]["cumulativeLTVByYear"] if years else 0
    company_bep_value = None
    annual_fixed_profit_for_bep = (
        scenario["salePrice"] / scenario["contractYears"] + scenario["annualFee"] + scenario["annualFixedProfit"]
    )
    if weighted_margin < 0:
        company_bep_value = math.floor(annual_fixed_profit_for_bep / abs(weighted_margin)) if annual_fixed_profit_for_bep > 0 else 0
    customer_bep_value = (
        math.ceil((scenario["salePrice"] + scenario["annualFee"] * scenario["contractYears"]) / weighted_saving)
        if weighted_saving > 0
        else math.inf
    )

    region_data = {}
    for region in scenario["regions"]:
        region_data[region["id"]] = {
            "region": region,
            "avgUsageRate": get_average_usage_rate_by_region(scenario, region),
            "actualRoundsByRegion": actual_rounds_by_region(scenario, region),
            "weightByRegion": (actual_rounds_by_region(scenario, region) / total_actual_rounds_value * 100)
            if total_actual_rounds_value
            else 0,
            "memberProfitByRegion": member_profit_by_region(region),
            "companionProfitByRegion": companion_profit_by_region(region),
            "marginPerRoundByRegion": margin_per_round_by_region(region),
            "annualRoundingByRegion": margin_per_round_by_region(region) * actual_rounds_by_region(scenario, region),
            "memberSavingByRegion": member_saving_by_region(region),
            "companionSavingByRegion": companion_saving_by_region(region),
            "savingPerRoundByRegion": saving_per_round_by_region(region),
            "annualSavingByRegion": saving_per_round_by_region(region) * actual_rounds_by_region(scenario, region),
            "peoplePerRoundByRegion": region["membersByRegion"] + region["companionsByRegion"],
            "membersPerRoundByRegion": region["membersByRegion"],
            "companionsPerRoundByRegion": region["companionsByRegion"],
        }

    return {
        "scenario": scenario,
        "years": years,
        "regionData": region_data,
        "totalRoundLimit": total_round_limit_value,
        "weightedUsageRate": sum(
            get_average_usage_rate_by_region(scenario, region) * (region["roundLimitByRegion"] / total_round_limit_value)
            for region in scenario["regions"]
        ),
        "weightedCourseCost": weighted_average_by_actual_rounds(scenario, lambda region: region["courseCostByRegion"]),
        "weightedMemberPrice": weighted_average_by_actual_rounds(scenario, lambda region: region["memberPriceByRegion"]),
        "weightedMembers": weighted_average_by_actual_rounds(scenario, lambda region: region["membersByRegion"]),
        "weightedCompanionPrice": weighted_average_by_actual_rounds(scenario, lambda region: region["companionPriceByRegion"]),
        "weightedCompanions": weighted_average_by_actual_rounds(scenario, lambda region: region["companionsByRegion"]),
        "weightedPeoplePerRound": weighted_average_by_actual_rounds(
            scenario, lambda region: region["membersByRegion"] + region["companionsByRegion"]
        ),
        "weightedPublicPrice": weighted_average_by_actual_rounds(scenario, lambda region: region["publicPriceByRegion"]),
        "sumActualRounds": total_actual_rounds_value,
        "weightedMarginPerRound": weighted_margin,
        "weightedMemberProfit": weighted_average_by_actual_rounds(scenario, member_profit_by_region),
        "weightedCompanionProfit": weighted_average_by_actual_rounds(scenario, companion_profit_by_region),
        "weightedMemberSaving": weighted_average_by_actual_rounds(scenario, member_saving_by_region),
        "weightedCompanionSaving": weighted_average_by_actual_rounds(scenario, companion_saving_by_region),
        "totalAnnualRounding": total_annual_rounding(scenario),
        "weightedSavingPerRound": weighted_saving,
        "totalAnnualSaving": total_annual_saving(scenario),
        "avgChurnRate": average(scenario.get("churnByYear", []), float(scenario.get("yearlyChurnRate", 0) or 0)),
        "avgRefundAmount": average(
            scenario.get("refundAmountByYear", []), float(scenario.get("yearlyRefundAmount", 0) or 0)
        ),
        "naiveFixedProfit": naive_fixed_profit_value,
        "ltv": ltv_value,
        "companyBEP": company_bep_value,
        "customerBEP": customer_bep_value,
        "naiveCustomerFixedCost": scenario["salePrice"] + scenario["annualFee"] * scenario["contractYears"],
        "avgAnnualProfit": ltv_value / scenario["contractYears"],
        "avgAnnualRounds": total_actual_rounds_value,
        "customerNetProfit": weighted_saving * total_actual_rounds_value * scenario["contractYears"]
        - scenario["salePrice"]
        - scenario["annualFee"] * scenario["contractYears"],
        "totalExpectedAnnualFeeRevenue": total_expected_annual_fee_revenue,
        "totalExpectedPositiveAnnualFixedProfit": total_expected_positive_annual_fixed_profit,
        "totalExpectedNegativeAnnualFixedCost": total_expected_negative_annual_fixed_cost,
        "totalExpectedRoundingRevenue": total_expected_rounding_revenue,
        "totalExpectedRoundingCost": total_expected_rounding_cost,
        "totalExpectedVariableRevenue": total_expected_variable_revenue,
        "totalExpectedVariableCost": total_expected_variable_cost,
        "avgAnnualVariableRevenue": avg_annual_variable_revenue,
        "avgAnnualVariableCost": avg_annual_variable_cost,
        "avgAnnualVariableProfit": avg_annual_variable_profit,
        "totalExpectedRefundCost": total_expected_refund_cost,
        "expectedEnrollProfit": scenario["salePrice"] - total_expected_refund_cost,
        "expectedMaturitySurvivingRate": years[-1]["survivingEndByYear"] if years else 0,
    }


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
        return None
    return value


def make_row(*, category: str, variable_key: str, value, meta: dict[str, str], region_name: str = "", year_index: int | None = None) -> dict:
    return {
        "분류": category,
        "변수명": variable_key,
        "라벨": meta.get("label", variable_key),
        "구장명": region_name or "",
        "연차": year_index if year_index is not None else "",
        "값": clean_value(value),
        "단위": meta.get("unit", ""),
        "tip": meta.get("tip", ""),
    }


def append_sheet_rows(sheet, rows: list[dict]) -> None:
    headers = ["분류", "변수명", "라벨", "구장명", "연차", "값", "단위", "tip"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row[column] for column in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 16,
        "B": 32,
        "C": 24,
        "D": 18,
        "E": 8,
        "F": 14,
        "G": 10,
        "H": 56,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = "0.0########"


def build_rows_for_scenario(scenario: dict, registry_meta: dict[str, dict[str, str]]) -> list[dict]:
    result = build_result_bundle(scenario)
    rows: list[dict] = []

    scenario_input_keys = [
        "salePrice",
        "annualFee",
        "annualFixedProfit",
        "contractYears",
        "yearlyChurnRate",
        "yearlyRefundAmount",
    ]
    for key in scenario_input_keys:
        rows.append(
            make_row(
                category="입력변수",
                variable_key=key,
                value=scenario.get(key),
                meta=registry_meta.get(key, {}),
            )
        )

    region_input_keys = [
        "roundLimitByRegion",
        "courseCostByRegion",
        "memberPriceByRegion",
        "membersByRegion",
        "companionPriceByRegion",
        "companionsByRegion",
        "publicPriceByRegion",
    ]
    for region in scenario["regions"]:
        for key in region_input_keys:
            rows.append(
                make_row(
                    category="입력변수",
                    variable_key=key,
                    region_name=region["name"],
                    value=region.get(key),
                    meta=registry_meta.get(key, {}),
                )
            )

    total_keys = [
        "totalRoundLimit",
        "weightedUsageRate",
        "weightedCourseCost",
        "weightedMemberPrice",
        "weightedMembers",
        "weightedCompanionPrice",
        "weightedCompanions",
        "weightedPeoplePerRound",
        "weightedPublicPrice",
        "totalActualRounds",
        "weightedMarginPerRound",
        "weightedMemberProfit",
        "weightedCompanionProfit",
        "weightedMemberSaving",
        "weightedCompanionSaving",
        "totalAnnualRounding",
        "weightedSavingPerRound",
        "totalAnnualSaving",
        "avgChurnRate",
        "avgRefundAmount",
        "naiveFixedProfit",
        "ltv",
        "companyBEP",
        "customerBEP",
        "naiveCustomerFixedCost",
        "avgAnnualProfit",
        "avgAnnualRounds",
        "customerNetProfit",
        "totalExpectedAnnualFeeRevenue",
        "totalExpectedPositiveAnnualFixedProfit",
        "totalExpectedNegativeAnnualFixedCost",
        "totalExpectedRoundingRevenue",
        "totalExpectedRoundingCost",
        "totalExpectedVariableRevenue",
        "totalExpectedVariableCost",
        "avgAnnualVariableRevenue",
        "avgAnnualVariableCost",
        "avgAnnualVariableProfit",
        "totalExpectedRefundCost",
        "expectedEnrollProfit",
        "expectedMaturitySurvivingRate",
    ]
    for key in total_keys:
        rows.append(
            make_row(
                category="총집계",
                variable_key=key,
                value=result.get(key),
                meta=registry_meta.get(key, {}),
            )
        )

    region_metric_keys = [
        "avgUsageRateByRegion",
        "actualRoundsByRegion",
        "weightByRegion",
        "marginPerRoundByRegion",
        "memberProfitByRegion",
        "companionProfitByRegion",
        "annualRoundingByRegion",
        "memberSavingByRegion",
        "companionSavingByRegion",
        "savingPerRoundByRegion",
        "annualSavingByRegion",
        "peoplePerRoundByRegion",
        "membersPerRoundByRegion",
        "companionsPerRoundByRegion",
    ]
    for region_entry in result["regionData"].values():
        region = region_entry["region"]
        for key in region_metric_keys:
            rows.append(
                make_row(
                    category="구장별 집계",
                    variable_key=key,
                    region_name=region["name"],
                    value=region_entry.get(key),
                    meta=registry_meta.get(key, {}),
                )
            )

    year_key_map = [
        ("yearSurvivingStartRate", "survivingStartByYear"),
        ("yearWeightedUsageRate", "usageByYear"),
        ("yearChurnRate", "rawChurnRate"),
        ("yearRefundAmount", "rawRefundAmount"),
        ("yearEnrollFee", "enrollFeeByYear"),
        ("annualFee", "annualFeeRevenueByYear"),
        ("annualFixedProfit", "annualFixedProfitByYear"),
        ("yearActualRoundsExpected", "actualRoundsExpectedByYear"),
        ("yearMemberRevenueExpected", "memberRevenueExpectedByYear"),
        ("yearCompanionRevenueExpected", "companionRevenueExpectedByYear"),
        ("yearRoundingRevenueExpected", "roundingRevenueExpectedByYear"),
        ("yearRoundingCostExpected", "roundingCostExpectedByYear"),
        ("yearRefundCost", "refundCostByYear"),
        ("yearVariableProfitExpected", "variableProfitExpectedByYear"),
        ("yearCashFlowExpected", "cashFlowExpectedByYear"),
        ("yearCumulativeLtv", "cumulativeLTVByYear"),
    ]
    for year in result["years"]:
        for variable_key, field in year_key_map:
            rows.append(
                make_row(
                    category="연도별 집계",
                    variable_key=variable_key,
                    year_index=year["yearByYear"],
                    value=year.get(field),
                    meta=registry_meta.get(variable_key, registry_meta.get("yearCumulativeLtv", {})),
                )
            )

    rows.extend(build_region_year_rows(scenario, registry_meta))
    return rows


def build_workbook() -> Workbook:
    registry_meta = parse_registry_meta()
    scenarios = load_default_scenarios()
    workbook = Workbook()
    workbook.remove(workbook.active)

    for scenario in scenarios:
        sheet = workbook.create_sheet(title=scenario["name"][:31])
        append_sheet_rows(sheet, build_rows_for_scenario(scenario, registry_meta))

    return workbook


def main() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
